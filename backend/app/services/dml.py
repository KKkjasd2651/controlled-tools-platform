"""DML.xlsx 操作 —— 严格使用用户提供的 TDML 模板。

模板有两个 sheet:
  List —— 工令固定的"应受控文档清单"（F1 = 工令号；A 列编码由公式拼成）
  Log  —— 实际受控履历，每次提交新版本就在第一个空白行追加 (日期, 文档编码, 版本)

List 中其他列（首版/末版受控、版本、链接等）通过数组公式 + VLOOKUP 自动反算。
"""
from __future__ import annotations

import shutil
from datetime import datetime
from pathlib import Path
from typing import Optional

from contextlib import closing
from datetime import date

from openpyxl import load_workbook
from openpyxl.worksheet.formula import ArrayFormula
from openpyxl.worksheet.worksheet import Worksheet


LIST_SHEET = "List"
LOG_SHEET = "Log"
PROJECT_CELL = "F1"   # 工令号
ROOT_DIR_CELL = "H1"  # 文件目录 (HYPERLINK 公式)

# 模板内 List 表 C/D/E/H 列的公式 —— append 新行时按行号注入
# 按 SBPC 现场约定，唯一键 = (文档编码, 文档标题)：同编码不同标题视为不同图。
# Log 表里搜该 (编码, 标题) 的所有版本号，取最大整数部分 N
# LOOKUP(2, 1/((A=a)*(B=b)), result) 是"复合条件最后匹配"惯用法。
LIST_FORMULA_C = (  # 版本 (Log 表里该 编码+标题 的最后一次写入)
    '=IFERROR(LOOKUP(2,1/'
    '((Log[文档编码]=List[[#This Row],[文档编码]])*(Log[文档标题]=List[[#This Row],[文档标题]])),'
    'Log[版本]),"")'
)
LIST_FORMULA_D = (  # 首版受控 - 数组公式（复合键）
    '=IFERROR(MIN(IF('
    '(Log[文档编码]=List[[#This Row],[文档编码]])*(Log[文档标题]=List[[#This Row],[文档标题]]),'
    'Log[日期])),"")'
)
LIST_FORMULA_E = (  # 末版受控 - 数组公式（复合键）
    '=IFERROR(MAX(IF('
    '(Log[文档编码]=List[[#This Row],[文档编码]])*(Log[文档标题]=List[[#This Row],[文档标题]]),'
    'Log[日期])),"")'
)
# H 列链接：反查 Log 表"完整文件名"列拿到真实落盘文件名（含版本+标题+扩展名），
# 拼到 $H$1 (工令文件夹基准 URL) 后面 → 用户点击可直接打开
# 例如：\\sbpc-dc.com\sbpc\项目\ceshi\260479\260455-C45H8-EF-15_A01 测试用.pdf
LIST_FORMULA_H = (
    '=IFERROR(HYPERLINK($H$1&LOOKUP(2,1/('
    '(Log[文档编码]=List[[#This Row],[文档编码]])*(Log[文档标题]=List[[#This Row],[文档标题]])),'
    'Log[完整文件名])),"")'
)

# 列单元格格式
NUMFMT_DATE = "yyyy/mm/dd"  # 与 TDML 模板的首版/末版受控列保持一致（带前导零）

# Log 表列索引（1-based，按模板顺序）
LOG_COL_C1 = 1            # =日期&文档编码
LOG_COL_C2 = 2            # =文档编码&版本
LOG_COL_DATE = 3          # 日期
LOG_COL_CODE = 4          # 文档编码
LOG_COL_VERSION = 5       # 版本
LOG_COL_TITLE = 6         # 文档标题 (VLOOKUP)
LOG_COL_FILENAME = 7      # 完整文件名 (拼接)
LOG_COL_LINK = 8          # 链接 (HYPERLINK)
LOG_COL_MD_CMD = 9        # 创建文件夹
LOG_COL_MV_CMD = 10       # 复制文件


def template_path(backend_root: Path) -> Path:
    return backend_root / "templates" / "DML_template.xlsx"


def _is_tdml_compatible(path: Path) -> bool:
    """检查现有 xlsx 是否符合 TDML 模板的工作表结构（含 List 和 Log）。"""
    try:
        with closing(load_workbook(path, read_only=True)) as wb:
            return LIST_SHEET in wb.sheetnames and LOG_SHEET in wb.sheetnames
    except Exception:
        return False


def ensure_project_dml(
    template: Path,
    target: Path,
    project_no: str,
    fs_root_url: Optional[str] = None,
) -> Path:
    """初始化或修复某工令的 DML 文件。

    三种分支：
      A. target 不存在 → 从模板复制 + 写工令号
      B. target 已存在且包含 TDML 必需工作表 (List + Log) → 不动
      C. target 已存在但缺少必需工作表 (例如 SBPC 现场的旧版"文件清单.xlsx")
         → 把它改名为 `文件清单_备份_YYYYMMDD-HHMMSS.xlsx` 保留，再从模板复制
    """
    import datetime as _dt
    target.parent.mkdir(parents=True, exist_ok=True)

    if not target.exists():
        shutil.copy2(template, target)
    elif not _is_tdml_compatible(target):
        # 备份不兼容的旧文件，再用模板覆盖
        ts = _dt.datetime.now().strftime("%Y%m%d-%H%M%S")
        backup = target.with_name(f"{target.stem}_备份_{ts}{target.suffix}")

        # 1) 先尝试 rename (最便宜)
        renamed = False
        try:
            target.rename(backup)
            renamed = True
        except PermissionError:
            pass  # 锁住了，下面尝试复制替代

        # 2) rename 失败 → 尝试复制源到备份位置 (只需读权限)
        if not renamed:
            try:
                shutil.copy2(target, backup)
            except PermissionError as e:
                raise RuntimeError(
                    f"无法备份旧的 {target.name}（{e}）—— 读权限也被独占。"
                    f"该文件被其他进程占用 (Excel / Explorer 预览 / 同步客户端 / SMB 服务端会话)。"
                    f"应对: a) 关闭所有 Excel 与该工令文件夹的 Explorer 窗口；"
                    f"b) 任务管理器结束 explorer.exe 再恢复；"
                    f"c) 手动把 {target} 改名后再上传；"
                    f"d) 重启电脑后再试。"
                ) from e

        # 3) 用模板覆盖目标位置
        try:
            shutil.copy2(template, target)
        except PermissionError as e:
            raise RuntimeError(
                f"已备份旧文件到 {backup.name}，但无法用模板覆盖 {target.name}（{e}）。"
                f"请关闭占用 {target} 的程序后重试。"
            ) from e
    else:
        # 已是 TDML 兼容格式，无须改动
        return target

    # 新拷贝的模板：写工令号 & 链接基准
    with closing(load_workbook(target)) as wb:
        lst: Worksheet = wb[LIST_SHEET]
        lst[PROJECT_CELL].value = project_no
        if fs_root_url:
            lst[ROOT_DIR_CELL].value = f'=HYPERLINK("{fs_root_url}","{fs_root_url}")'
        wb.save(target)
    return target


def read_project_no(path: Path) -> str:
    with closing(load_workbook(path, read_only=True, data_only=False)) as wb:
        return str(wb[LIST_SHEET][PROJECT_CELL].value or "")


def list_codes(path: Path) -> list[str]:
    """返回 List 表所有文档编码（合并 data_only 缓存值 + 公式回退）。

    合并而非二选一，避免：data_only 拿到 append 行的纯字符串（少量）→
    误以为 data_only 已就绪 → 漏掉公式行（如预设的 47 条）。
    """
    codes: set[str] = set()

    # 1) data_only 拿已缓存的字符串（包括之前 append 的、Excel 算过的公式值）
    try:
        with closing(load_workbook(path, data_only=True, read_only=True)) as wb_calc:
            if LIST_SHEET in wb_calc.sheetnames:
                lst = wb_calc[LIST_SHEET]
                for r in range(5, lst.max_row + 1):
                    v = lst.cell(r, 1).value
                    if v:
                        codes.add(str(v).strip())
    except Exception:
        pass

    # 2) 公式回退：模板里的 `=$F$1&"-XX-XX-NNNN"` 自行拼接
    try:
        with closing(load_workbook(path, data_only=False, read_only=True)) as wb:
            if LIST_SHEET in wb.sheetnames:
                lst = wb[LIST_SHEET]
                project = str(lst[PROJECT_CELL].value or "")
                for r in range(5, lst.max_row + 1):
                    v = lst.cell(r, 1).value
                    if not isinstance(v, str):
                        continue
                    if v.startswith("=$F$1&"):
                        try:
                            suffix = v.split('"', 2)[1]
                            codes.add(f"{project}{suffix}")
                        except IndexError:
                            pass
                    else:
                        codes.add(v.strip())
    except Exception:
        pass

    return sorted(c for c in codes if c)


def log_rows(path: Path) -> list[dict]:
    """读取 Log 中的实际数据行；Log 工作表不存在时返回空列表。"""
    out: list[dict] = []
    try:
        with closing(load_workbook(path, data_only=False, read_only=True)) as wb:
            if LOG_SHEET not in wb.sheetnames:
                return []
            log = wb[LOG_SHEET]
            for r in range(2, log.max_row + 1):
                date_v = log.cell(r, LOG_COL_DATE).value
                code_v = log.cell(r, LOG_COL_CODE).value
                if date_v in (None, "") and code_v in (None, ""):
                    continue
                out.append({
                    "row": r,
                    "日期": date_v.strftime("%Y-%m-%d") if hasattr(date_v, "strftime") else date_v,
                    "文档编码": code_v,
                    "版本": log.cell(r, LOG_COL_VERSION).value,
                })
    except Exception:
        return []
    return out


def find_existing_version(path: Path, code: str, version: str) -> bool:
    for r in log_rows(path):
        if str(r["文档编码"]) == code and str(r["版本"]) == version:
            return True
    return False


def append_log_entry(
    path: Path,
    code: str,
    version: str,
    title: Optional[str] = None,
    filename: Optional[str] = None,
    when: Optional[datetime] = None,
) -> int:
    """在 Log 表第一个空白行写入 (日期, 文档编码, 版本, 文档标题, 完整文件名)。

    - 日期只取到"日"（不带时分秒），避免 List 反查显示成 46156.3895 序列号。
    - 文档标题列：原本是 VLOOKUP 公式（按编码反查 List），改写为字符串，
      让 List 表能按 (编码, 标题) 复合键正确反查。
    - 完整文件名列：原本是 `=编码&"_"&版本&" "&标题` 公式（不含扩展名），
      改写为用户上传时的**真实原文件名**（含扩展名），让 List H 列 HYPERLINK
      能拼出可点击的真实文件路径。
    """
    when = when or datetime.now()
    day_only = date(when.year, when.month, when.day)

    with closing(load_workbook(path)) as wb:
        log = wb[LOG_SHEET]

        target_row = None
        for r in range(2, log.max_row + 1):
            if (
                log.cell(r, LOG_COL_DATE).value in (None, "")
                and log.cell(r, LOG_COL_CODE).value in (None, "")
            ):
                target_row = r
                break
        if target_row is None:
            target_row = log.max_row + 1

        log.cell(target_row, LOG_COL_DATE).value = day_only
        log.cell(target_row, LOG_COL_DATE).number_format = NUMFMT_DATE
        log.cell(target_row, LOG_COL_CODE).value = code
        log.cell(target_row, LOG_COL_VERSION).value = version
        if title:
            log.cell(target_row, LOG_COL_TITLE).value = title
        if filename:
            log.cell(target_row, LOG_COL_FILENAME).value = filename

        wb.save(path)
    return target_row


def next_log_version(path: Path, code: str) -> str:
    """计算某文档编码在 Log 表里的下一个版本号。

    规则：
      - Log 表里搜该编码的所有版本号，取最大整数部分 N
      - 返回 "{N+1}.0"
      - 没有任何已存在记录 → "1.0"

    例：原有 1.0、1.5、2.0 → 下一个 3.0
       原有 1.0 → 下一个 2.0
       完全新编码 → 1.0
    """
    max_int = 0
    for r in log_rows(path):
        if str(r.get("文档编码", "")).strip() == code:
            v = str(r.get("版本", "")).strip()
            try:
                # 允许 "1.0" / "1" / "1.5" 等格式
                num = float(v)
                if int(num) > max_int:
                    max_int = int(num)
            except (ValueError, TypeError):
                continue
    return f"{max_int + 1}.0"


# List 表列索引（1-based，按模板顺序）
LIST_COL_CODE = 1     # 文档编码 (A)
LIST_COL_TITLE = 2    # 文档标题 (B)
LIST_COL_AUTHOR = 6   # 作者 (F)


def _find_code_title_row(path: Path, code: str, title: Optional[str]) -> Optional[int]:
    """按 (code, title) 复合主键定位 List 行号；找不到返回 None。

    匹配规则：
      - A 列（文档编码）必须 == code（覆盖 data_only 缓存 / 公式拼接 / 纯字符串三种形态）
      - 同时 B 列（文档标题）必须 == title
      - 例外：若 title 为空（用户未提供），只按 code 匹配，找到第一个就返回

    覆盖三种 A 列形态：
      ① data_only 已缓存的字符串（Excel 算过的预设编码）
      ② 模板里的公式 `=$F$1&"-XX-XX-NNNN"`（新建 DML 时尚未算过）
      ③ 之前 append 时写入的纯字符串
    """
    title_norm = (title or "").strip()

    def _match_title(row_title) -> bool:
        if not title_norm:
            return True  # title 未提供 → 只比 code
        return str(row_title or "").strip() == title_norm

    # ① data_only
    try:
        with closing(load_workbook(path, data_only=True, read_only=True)) as wb_calc:
            if LIST_SHEET in wb_calc.sheetnames:
                lst_calc = wb_calc[LIST_SHEET]
                for r in range(5, lst_calc.max_row + 1):
                    v = lst_calc.cell(r, LIST_COL_CODE).value
                    if v and str(v).strip() == code:
                        if _match_title(lst_calc.cell(r, LIST_COL_TITLE).value):
                            return r
    except Exception:
        pass

    # ② / ③ 公式 fallback + 纯字符串
    try:
        with closing(load_workbook(path, data_only=False, read_only=True)) as wb:
            if LIST_SHEET not in wb.sheetnames:
                return None
            lst = wb[LIST_SHEET]
            project_no = str(lst[PROJECT_CELL].value or "")
            for r in range(5, lst.max_row + 1):
                v = lst.cell(r, LIST_COL_CODE).value
                code_match = False
                if isinstance(v, str):
                    if v.startswith("=$F$1&"):
                        try:
                            suffix = v.split('"', 2)[1]
                        except IndexError:
                            continue
                        code_match = (f"{project_no}{suffix}" == code)
                    elif v == code:
                        code_match = True
                if code_match and _match_title(lst.cell(r, LIST_COL_TITLE).value):
                    return r
    except Exception:
        pass
    return None


def _ensure_list_formulas(lst: Worksheet, row: int) -> None:
    """确保 List 表某行的 C/D/E/H 公式列正确 + D/E 列日期格式。

    - **C 列强制覆盖**为 LOOKUP "最后匹配" 公式：TDML 原模板用 VLOOKUP 在
      同一天多次升版本时取不到最新版（截图里 Log 有 V1.0/2.0/3.0/4.0 但
      List 只显示 V2.0 就是这个原因）。每次 upsert 都把它改写为新公式，
      历史行也借此被修复。
    - D/E 公式只在为空时写入，但 number_format **每次都强制刷新**为日期。
    - H 公式为空才写入。
    """
    # C - 版本（强制覆盖为 LOOKUP）
    lst.cell(row, 3).value = LIST_FORMULA_C
    # D - 首版受控
    if lst.cell(row, 4).value in (None, ""):
        lst.cell(row, 4).value = ArrayFormula(f"D{row}", LIST_FORMULA_D)
    lst.cell(row, 4).number_format = NUMFMT_DATE
    # E - 末版受控
    if lst.cell(row, 5).value in (None, ""):
        lst.cell(row, 5).value = ArrayFormula(f"E{row}", LIST_FORMULA_E)
    lst.cell(row, 5).number_format = NUMFMT_DATE
    # H - 链接
    if lst.cell(row, 8).value in (None, ""):
        lst.cell(row, 8).value = LIST_FORMULA_H


def _extend_list_table_ref(lst: Worksheet, new_max_row: int) -> None:
    """把 List 表的 Excel Table 范围扩展到 new_max_row。

    Excel Table 范围（ref）不会随着写入行自动扩展；如果新行不在 Table 内，
    `List[[#This Row],[...]]` 这种结构化引用就会失效。
    """
    try:
        table = lst.tables.get(LIST_SHEET)
        if table is None:
            return
        ref = table.ref  # 例如 "A4:M55"
        start, _, end = ref.partition(":")
        # end 形如 "M55"
        end_col = "".join(ch for ch in end if ch.isalpha())
        end_row = int("".join(ch for ch in end if ch.isdigit()) or "0")
        if new_max_row > end_row:
            table.ref = f"{start}:{end_col}{new_max_row}"
    except Exception:
        # Table 扩展失败不阻断主流程；至少 A/B/F 列的明文值是正确的
        pass


def upsert_list_row(
    path: Path,
    code: str,
    title: Optional[str],
    author: Optional[str],
) -> tuple[int, str]:
    """在 List 表登记某编码的标题/作者。

    行为：
      - 编码已存在（命中预设 47 条之一 或 之前追加过）→ 更新该行的 B (标题) 和 F (作者)；
        若该行缺少 C/D/E/H 公式，顺便补上
      - 编码不存在 → 在 List 表底部追加一行：
            A=编码字符串、B=标题、F=作者、C/D/E/H=反查公式
        并把 Excel Table 范围扩展到新行号

    返回：(行号, 动作)，动作 ∈ {"update", "append"}
    """
    target_row = _find_code_title_row(path, code, title)

    # 再用 data_only=False 实际写入（避免破坏公式列）
    with closing(load_workbook(path)) as wb:
        lst = wb[LIST_SHEET]

        if target_row is not None:
            action = "update"
            # 命中预设行：B 列原本可能是硬编码标题，按需更新；F 列原本可能空，写入作者
            if title:
                lst.cell(target_row, LIST_COL_TITLE).value = title
            if author:
                lst.cell(target_row, LIST_COL_AUTHOR).value = author
            _ensure_list_formulas(lst, target_row)
        else:
            action = "append"
            target_row = lst.max_row + 1
            lst.cell(target_row, LIST_COL_CODE).value = code
            if title:
                lst.cell(target_row, LIST_COL_TITLE).value = title
            if author:
                lst.cell(target_row, LIST_COL_AUTHOR).value = author
            _ensure_list_formulas(lst, target_row)
            _extend_list_table_ref(lst, target_row)

        # 一次性刷新所有非空 List 行 —— 借机修复历史行的旧 VLOOKUP 公式和日期格式
        for row in range(5, lst.max_row + 1):
            if row == target_row:
                continue
            if lst.cell(row, 1).value or lst.cell(row, 2).value:
                _ensure_list_formulas(lst, row)

        wb.save(path)
    return target_row, action
